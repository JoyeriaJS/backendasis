from fastapi import FastAPI, Depends
from datetime import datetime, date
from sqlalchemy.orm import Session
from math import radians, cos, sin, asin, sqrt
from fastapi.middleware.cors import CORSMiddleware

from app.database import SessionLocal, engine, Base
from app.models import Attendance, User, Documento
from fastapi.responses import FileResponse
from openpyxl import Workbook
import os
from fastapi import HTTPException
from zoneinfo import ZoneInfo

from uuid import uuid4
from fastapi import UploadFile, File, Form
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi.responses import FileResponse
from datetime import datetime, timedelta
from collections import defaultdict
import calendar
import pyotp
import qrcode
from io import BytesIO
from fastapi.responses import StreamingResponse
from app.mail_service import enviar_comprobante


app = FastAPI()

# ✅ CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ✅ crear tablas
Base.metadata.create_all(bind=engine)

# 📍 ubicación empresa (TU TIENDA)
LAT_EMPRESA = -33.43943
LNG_EMPRESA = -70.648964
DISTANCIA_MAX_METROS = 100


# ✅ conexión BD
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ✅ fórmula distancia (metros)
def calcular_distancia_metros(lat1, lon1, lat2, lon2):
    r = 6371000  # metros
    dlat = radians(lat2 - lat1)
    dlon = radians(lon2 - lon1)

    a = sin(dlat/2)**2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a))

    return r * c


# 🔐 LOGIN
@app.post("/login")
def login(data: dict, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == data["username"]).first()

    if not user or user.password != data["password"]:
        return {"error": "Credenciales inválidas"}

    return {
        "user_id": user.id,
        "username": user.username,
        "role": user.role
    }


# 📍 MARCAR ASISTENCIA REAL
@app.post("/marcar")
def marcar(data: dict, db: Session = Depends(get_db)):

    # ✅ validar datos
    if "user_id" not in data:
        return {"error": "Falta user_id"}

    if "lat" not in data or "lng" not in data:
        return {"error": "Faltan coordenadas"}

    user_id = data["user_id"]
    codigo = data.get("codigo")

    user = db.query(User).filter(User.id == user_id).first()

    if not user:
        return {"error": "Usuario no encontrado"}

    if not user.totp_secret:
        return {"error": "Authenticator no configurado"}

    totp = pyotp.TOTP(user.totp_secret)

    if not totp.verify(codigo):

        return {
            "error": "Código Authenticator inválido"
        }
    lat = data["lat"]
    lng = data["lng"]
    
    # 👇 viene del frontend
    accuracy = data.get("accuracy", 999)

    # 📍 calcular distancia
    distancia = calcular_distancia_metros(lat, lng, LAT_EMPRESA, LNG_EMPRESA)

    # 🔥 SOLUCIÓN REAL PARA INTERIOR (galería, piso 9, etc)
    if accuracy >= 1000:
        # GPS muy malo → permitir pero con límite razonable
        margen = 600   # 👈 ajusta entre 400 y 800 según pruebas
    else:
        # GPS decente
        margen = DISTANCIA_MAX_METROS + (accuracy * 0.7)

    # 🚫 validación final (UNA sola vez)
    if distancia > margen:
        return {
            "error": "Fuera de la zona permitida",
            "distancia_metros": round(distancia, 2),
            "accuracy": accuracy,
            "margen": margen
        }

    # 📅 registros de HOY
    hoy = datetime.now(ZoneInfo("America/Santiago")).date()

    registros_hoy = (
        db.query(Attendance)
        .filter(
            Attendance.user_id == user_id,
            Attendance.fecha >= datetime.combine(hoy, datetime.min.time())
        )
        .order_by(Attendance.fecha.asc())
        .all()
    )

    if len(registros_hoy) >= 4:
        return {"error": "Ya completaste tus 4 marcajes del día"}

    tipos = [
        "entrada",
        "salida_colacion",
        "entrada_colacion",
        "salida"
    ]

    tipo = tipos[len(registros_hoy)]

    nuevo = Attendance(
        user_id=user_id,
        username=user.username,
        lat=lat,
        lng=lng,
        fecha=datetime.now(ZoneInfo("America/Santiago")).replace(tzinfo=None),
        tipo=tipo
    )

    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)

    return {
        "message": f"{tipo.replace('_', ' ').capitalize()} registrada",
        "tipo": tipo,
        "distancia_metros": round(distancia, 2),
        "accuracy": accuracy,
        "margen": margen
    }

@app.get("/estado/{user_id}")
def estado(user_id: int, db: Session = Depends(get_db)):
    from datetime import date

    hoy = date.today()

    registros = (
        db.query(Attendance)
        .filter(
            Attendance.user_id == user_id,
            Attendance.fecha >= datetime.combine(hoy, datetime.min.time())
        )
        .order_by(Attendance.fecha.asc())
        .all()
    )

    tipos = ["Entrada", "Salida a colación", "Entrada de colación", "Salida"]

    if len(registros) >= 4:
        return {"estado": "completo", "siguiente": None}

    return {
        "estado": "pendiente",
        "siguiente": tipos[len(registros)]
    }
@app.get("/asistencia/{user_id}")
def asistencia(user_id: int, db: Session = Depends(get_db)):
    return db.query(Attendance)\
        .filter(Attendance.user_id == user_id)\
        .order_by(Attendance.fecha.desc())\
        .all()




from fastapi import Request
from datetime import datetime, timedelta

@app.get("/historial")
def historial(
    user_id: int,
    fecha_inicio: str = None,
    fecha_fin: str = None,
    db: Session = Depends(get_db)
):
    query = db.query(Attendance).filter(Attendance.user_id == user_id)

    if fecha_inicio:
        fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        query = query.filter(Attendance.fecha >= fi)

    if fecha_fin:
        ff = datetime.strptime(fecha_fin, "%Y-%m-%d")
        ff = ff + timedelta(days=1)  # 🔥 CLAVE
        query = query.filter(Attendance.fecha < ff)

    registros = query.order_by(Attendance.fecha.desc()).all()

    return registros

from collections import defaultdict
from datetime import datetime

from collections import defaultdict
from datetime import datetime


from fastapi import Query

@app.get("/resumen")
def resumen(
    user_id: int = Query(None),
    fecha_inicio: str = Query(None),
    fecha_fin: str = Query(None),
    db: Session = Depends(get_db)
):

    query = db.query(Attendance)

    # 🔍 filtro por usuario
    if user_id:
        query = query.filter(Attendance.user_id == user_id)

    # 🔍 filtro por fechas
    if fecha_inicio:
        query = query.filter(Attendance.fecha >= fecha_inicio)

    if fecha_fin:
        query = query.filter(Attendance.fecha <= fecha_fin)

    registros = query.order_by(Attendance.fecha).all()
    users = db.query(User).all()

    from collections import defaultdict

    resumen = defaultdict(lambda: {
        "user": "",
        "horas": 0,
        "atrasos": 0,
        "retiros": 0
    })

    agrupado = defaultdict(lambda: defaultdict(list))

    for r in registros:
        fecha = r.fecha.date()
        agrupado[r.user_id][fecha].append(r)

    for user_id, dias in agrupado.items():

        user = next((u for u in users if u.id == user_id), None)
        if not user:
            continue

        resumen[user_id]["user"] = user.username

        for fecha, registros_dia in dias.items():

            registros_dia.sort(key=lambda x: x.fecha)

            entrada = None
            salida = None
            salida_colacion = None
            entrada_colacion = None

            for r in registros_dia:

                hora = r.fecha.hour

                if r.tipo == "entrada" and not entrada:
                    entrada = r.fecha
                    if hora >= 10:
                        resumen[user_id]["atrasos"] += 1

                elif r.tipo == "salida":
                    salida = r.fecha
                    if hora < 18:
                        resumen[user_id]["retiros"] += 1

                elif r.tipo == "salida_colacion" and not salida_colacion:
                    salida_colacion = r.fecha

                elif r.tipo == "entrada_colacion":
                    entrada_colacion = r.fecha

            if entrada and salida and salida > entrada:

                total = (salida - entrada).total_seconds()

                if salida_colacion and entrada_colacion and entrada_colacion > salida_colacion:
                    colacion = (entrada_colacion - salida_colacion).total_seconds()
                    total -= colacion

                if total < 0:
                    total = 0

                resumen[user_id]["horas"] += total

    resultado = []

    for r in resumen.values():
        horas = int(r["horas"] // 3600)
        minutos = int((r["horas"] % 3600) // 60)

        resultado.append({
            "user_id": user_id,
            "user": r["user"],
            "horas": f"{horas}h {minutos}min",
            "atrasos": r["atrasos"],
            "retiros": r["retiros"]
        })

    return resultado

@app.get("/usuarios")
def obtener_usuarios(db: Session = Depends(get_db)):
    users = db.query(User).all()

    return [
        {
            "id": u.id,
            "username": u.username
        }
        for u in users
    ]


@app.get("/siguiente/{user_id}")
def siguiente(user_id: int, db: Session = Depends(get_db)):

    hoy = date.today()

    registros = db.query(Attendance)\
        .filter(Attendance.user_id == user_id)\
        .all()

    # filtrar solo hoy
    registros_hoy = [r for r in registros if r.fecha.date() == hoy]

    tipos = [r.tipo for r in registros_hoy]

    if "entrada" not in tipos:
        return {"tipo": "entrada"}

    if "salida_colacion" not in tipos:
        return {"tipo": "salida_colacion"}

    if "entrada_colacion" not in tipos:
        return {"tipo": "entrada_colacion"}

    if "salida" not in tipos:
        return {"tipo": "salida"}

    return {"tipo": "bloqueado"}

from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from collections import defaultdict

from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import os


@app.get("/exportar-excel")
def exportar_excel(
    user_id: int = Query(None),
    fecha_inicio: str = Query(None),
    fecha_fin: str = Query(None),
    db: Session = Depends(get_db)
):

    query = db.query(Attendance)

    if user_id:
        query = query.filter(Attendance.user_id == user_id)

    if fecha_inicio:
        fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        query = query.filter(Attendance.fecha >= fi)

    if fecha_fin:
        ff = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1)
        query = query.filter(Attendance.fecha < ff)

    registros = query.order_by(
        Attendance.fecha.asc()
    ).all()

    users = {
        u.id: u.username
        for u in db.query(User).all()
    }

    wb = Workbook()

    # 🔥 HOJA RESUMEN
    ws = wb.active
    ws.title = "Resumen RRHH"

    headers = [
        "Usuario",
        "Días trabajados",
        "Horas Totales",
        "Horas Promedio",
        "Atrasos",
        "Retiros Anticipados",
        "% Asistencia",
        "Horas Colación",
        "Horas Extras",
        "Días Incompletos"
    ]

    ws.append(headers)

    fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    for cell in ws[1]:
        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )
        cell.fill = fill

    agrupado = defaultdict(
        lambda: defaultdict(list)
    )

    # 🔥 CONVERTIR TODO A HORA CHILE
    for r in registros:

        fecha_chile = r.fecha.astimezone(
            ZoneInfo("America/Santiago")
        )

        agrupado[r.user_id][fecha_chile.date()].append({
            "tipo": r.tipo,
            "fecha": fecha_chile
        })

    # 🔥 FORMATO HORAS
    def format_horas(segundos):

        horas = int(segundos // 3600)

        minutos = int(
            (segundos % 3600) // 60
        )

        return f"{horas}h {minutos}m"

    # 🔥 HOJA DETALLE
    ws_detalle = wb.create_sheet(
        title="Marcajes"
    )

    detalle_headers = [
        "Usuario",
        "Fecha",
        "Entrada",
        "Salida Colación",
        "Entrada Colación",
        "Salida",
        "Horas Trabajadas"
    ]

    ws_detalle.append(detalle_headers)

    for cell in ws_detalle[1]:
        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )
        cell.fill = fill

    # 🔥 RESUMEN
    for uid, dias in agrupado.items():

        atrasos = 0
        retiros = 0
        horas_totales = 0
        horas_colacion_total = 0
        horas_extras = 0
        incompletos = 0
        dias_trabajados = 0

        for fecha, registros_dia in dias.items():

            registros_dia.sort(
                key=lambda x: x["fecha"]
            )

            entrada = None
            salida = None
            salida_colacion = None
            entrada_colacion = None

            for r in registros_dia:

                if r["tipo"] == "entrada":
                    entrada = r["fecha"]

                    if entrada.hour >= 10:
                        atrasos += 1

                elif r["tipo"] == "salida":
                    salida = r["fecha"]

                    if salida.hour < 18:
                        retiros += 1

                elif r["tipo"] == "salida_colacion":
                    salida_colacion = r["fecha"]

                elif r["tipo"] == "entrada_colacion":
                    entrada_colacion = r["fecha"]

            horas_dia = 0

            if entrada and salida:

                dias_trabajados += 1

                total = (
                    salida - entrada
                ).total_seconds()

                colacion_seg = 0

                if (
                    salida_colacion and
                    entrada_colacion
                ):

                    colacion_seg = (
                        entrada_colacion -
                        salida_colacion
                    ).total_seconds()

                    total -= colacion_seg

                horas_dia = total

                horas_totales += total
                horas_colacion_total += colacion_seg

                if total > (9 * 3600):

                    horas_extras += (
                        total - (9 * 3600)
                    )

            else:
                incompletos += 1

            # 🔥 AGREGAR MARCAJES
            ws_detalle.append([
                users.get(uid, f"User {uid}"),
                fecha.strftime("%Y-%m-%d"),

                entrada.strftime("%H:%M")
                if entrada else "-",

                salida_colacion.strftime("%H:%M")
                if salida_colacion else "-",

                entrada_colacion.strftime("%H:%M")
                if entrada_colacion else "-",

                salida.strftime("%H:%M")
                if salida else "-",

                format_horas(horas_dia)
            ])

        promedio = (
            horas_totales / dias_trabajados
            if dias_trabajados > 0 else 0
        )

        dias_laborales = 22

        asistencia = (
            (dias_trabajados / dias_laborales) * 100
            if dias_laborales > 0 else 0
        )

        ws.append([
            users.get(uid, f"User {uid}"),
            dias_trabajados,
            format_horas(horas_totales),
            format_horas(promedio),
            atrasos,
            retiros,
            f"{asistencia:.1f}%",
            format_horas(horas_colacion_total),
            format_horas(horas_extras),
            incompletos
        ])

    # 🔥 AUTO AJUSTE COLUMNAS
    for hoja in [ws, ws_detalle]:

        for col in hoja.columns:

            max_length = 0

            column = get_column_letter(
                col[0].column
            )

            for cell in col:

                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(
                            str(cell.value)
                        )
                except:
                    pass

            hoja.column_dimensions[
                column
            ].width = max_length + 5

    file_path = "reporte_rrhh.xlsx"

    wb.save(file_path)

    return FileResponse(
        path=file_path,
        filename="reporte_asistencia.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            "attachment; filename=reporte_asistencia.xlsx"
        }
    )
def require_admin(user):
    if user.rol != "admin":
        raise HTTPException(status_code=403, detail="No autorizado")
    

@app.get("/documentos")
def get_documentos(user_id: int, db: Session = Depends(get_db)):

    docs = db.query(Documento)\
        .filter(Documento.user_id == user_id)\
        .order_by(Documento.id.desc())\
        .all()

    return docs

from pydantic import BaseModel
from typing import Optional

class FirmaDocumento(BaseModel):
    doc_id: int
    aprobado: bool
    observacion: Optional[str] = None
    correo: Optional[str] = None

@app.post("/documentos/firmar")
def firmar_documento(
    data: FirmaDocumento,
    db: Session = Depends(get_db)
):

    doc = db.query(Documento).filter(
        Documento.id == data.doc_id
    ).first()

    if not doc:
        return {"error": "Documento no encontrado"}

    doc.estado = "firmado" if data.aprobado else "rechazado"

    doc.observacion = data.observacion

    doc.correo = data.correo

    doc.fecha_firma = datetime.now()

    db.commit()

    # 🔥 BUSCAR USUARIO
    usuario = db.query(User).filter(
        User.id == doc.user_id
    ).first()

    # 🔥 ENVIAR CORREO
    try:
        from zoneinfo import ZoneInfo
        fecha_chile = datetime.now(
            ZoneInfo("America/Santiago")
        ).strftime("%d/%m/%Y %H:%M")

        enviar_comprobante(
            destino=data.correo,
            usuario=usuario.username,
            documento=doc.tipo,
            estado=doc.estado,
            fecha=fecha_chile,
            observacion=doc.observacion,
            archivo_url=doc.archivo_url,
            nombre_archivo=doc.nombre
        )

        print("✅ CORREO ENVIADO")

    except Exception as e:

        print("❌ ERROR SMTP:")
        print(str(e))

    return {
        "message": "Documento actualizado"
    }

@app.get("/admin/documentos")
def admin_documentos(
    user_id: int = Query(None),
    estado: str = Query(None),
    fecha_inicio: str = Query(None),
    fecha_fin: str = Query(None),
    orden: str = Query("desc"),
    db: Session = Depends(get_db)
):

    query = db.query(Documento, User)\
        .join(User, Documento.user_id == User.id)

    # 🔍 FILTRO USUARIO
    if user_id:
        query = query.filter(Documento.user_id == user_id)

    # 🔍 FILTRO ESTADO
    if estado:
        query = query.filter(Documento.estado == estado)

    # 🔍 FECHAS
    if fecha_inicio:
        fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        query = query.filter(Documento.fecha_firma >= fi)

    if fecha_fin:
        ff = datetime.strptime(fecha_fin, "%Y-%m-%d")
        ff = ff + timedelta(days=1)
        query = query.filter(Documento.fecha_firma < ff)

    # 🔥 ORDEN
    if orden == "asc":
        query = query.order_by(Documento.id.asc())
    else:
        query = query.order_by(Documento.id.desc())

    resultados = query.all()

    return [
        {
            "id": doc.id,
            "usuario": user.username,
            "tipo": doc.tipo,
            "periodo": doc.periodo,
            "estado": doc.estado,
            "archivo_url": doc.archivo_url,
            "fecha_firma": doc.fecha_firma,
            "observacion": doc.observacion
        }
        for doc, user in resultados
    ]

from fastapi import UploadFile, File, Form
import shutil
import os

#UPLOAD_DIR = "uploads"
from app.supabase_client import supabase
@app.post("/documentos/subir")
async def subir_documento(
    file: UploadFile = File(...),
    user_id: int = Form(...),
    tipo: str = Form(...),
    periodo: str = Form(...),
    db: Session = Depends(get_db)
):

    try:

        # 🔥 VALIDACIÓN PDF
        if file.content_type != "application/pdf":
            return {
                "error": "Solo se permiten archivos PDF"
            }

        contenido = await file.read()

        nombre_unico = f"{uuid4()}_{file.filename}"

        resultado = supabase.storage.from_("documentos").upload(
            path=nombre_unico,
            file=contenido,
            file_options={
                "content-type": file.content_type
            }
        )

        url = supabase.storage.from_("documentos").get_public_url(nombre_unico)

        nuevo = Documento(
            user_id=user_id,
            tipo=tipo,
            periodo=periodo,
            archivo_url=url,
            estado="pendiente"
        )

        db.add(nuevo)
        db.commit()

        return {
            "message": "Documento subido correctamente",
            "url": url
        }

    except Exception as e:
        print("ERROR SUBIENDO:", str(e))
        return {"error": str(e)}


from fastapi.staticfiles import StaticFiles
import os

#os.makedirs("uploads", exist_ok=True)
#app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")

@app.get("/authenticator/setup/{user_id}")
def setup_authenticator(
    user_id: int,
    db: Session = Depends(get_db)
):

    user = db.query(User).filter(User.id == user_id).first()

    if not user:
        return {"error": "Usuario no encontrado"}

    # 🔥 generar secret si no existe
    if not user.totp_secret:

        user.totp_secret = pyotp.random_base32()

        db.commit()

    totp = pyotp.TOTP(user.totp_secret)

    uri = totp.provisioning_uri(
        name=user.username,
        issuer_name="Joyeria Sebastian"
    )

    return {
        "otpauth_url": uri
    }

@app.get("/authenticator/qr/{user_id}")
def authenticator_qr(
    user_id: int,
    db: Session = Depends(get_db)
):

    user = db.query(User).filter(User.id == user_id).first()

    if not user:
        return {"error": "Usuario no encontrado"}

    if not user.totp_secret:

        user.totp_secret = pyotp.random_base32()

        db.commit()

    totp = pyotp.TOTP(user.totp_secret)

    uri = totp.provisioning_uri(
        name=user.username,
        issuer_name="Joyeria Sebastian"
    )

    qr = qrcode.make(uri)

    buffer = BytesIO()

    qr.save(buffer, format="PNG")

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="image/png"
    )